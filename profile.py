import openpyxl

class ResumeGenerator:
    def __init__(self, excel_file, output_tex):
        self.excel_file = excel_file
        self.output_tex = output_tex
        self.data = {}
        self.load_data()

    def load_data(self):
        wb = openpyxl.load_workbook(self.excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            section_data = [cell.value for row in ws.iter_rows(values_only=True) for cell in row]
            self.data[sheet.lower()] = section_data

    def generate_profile(self):
        profile = f"""
        %-------------------------------------------------------------------------------
        % PROFILE
        %-------------------------------------------------------------------------------
        \\photo[rectangle,edge,right]{{./resume/lttsLogo.png}}\\First Name{{{self.data['profile'][0]}}}
        \\Last Name{{{self.data['profile'][1]}}}
        \\Designation (Job Title){{{self.data['profile'][2]}}}
        \\Mobile No{{{self.data['profile'][3]}}}
        \\Official E-Mail Id{{{self.data['profile'][4]}}}
        \\DOB:{{{self.data['profile'][5]}}}
        \\Github Profile Link (Optional){{{self.data['profile'][6]}}}
        """
        return profile

    def generate_tex(self):
        with open(self.output_tex, 'w') as f:
            f.write(r"""
            %-------------------------------------------------------------------------------
            % CONFIGURATIONS
            %-------------------------------------------------------------------------------
            \documentclass[11pt, a4paper]{awesome-cv} 
            \geometry{left=1.4cm, top=.8cm, right=1.4cm, bottom=1.8cm, footskip=.5cm}
            \colorlet{awesome}{awesome-skyblue}
            \setbool{acvSectionColorHighlight}{true}
            \renewcommand{\acvHeaderSocialSep}{\quad\textbar\quad}
            """)

            f.write(self.generate_profile())
            f.write(r"""
            \input{resume/summary.tex}
            \input{resume/skills.tex}
            \input{resume/competencies.tex}
            \input{resume/experience.tex}
            \input{resume/education.tex}

            \end{document}
            """)

if __name__ == "__main__":
    excel_file = "data.xlsx"  # Replace with your Excel file path
    output_tex = "main_resume.tex"
    
    generator = ResumeGenerator(excel_file, output_tex)
    generator.generate_tex()

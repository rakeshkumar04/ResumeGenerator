import openpyxl
import re
from datetime import datetime

class ResumeGenerator:
    def __init__(self, excel_file, output_tex):
        self.excel_file = excel_file
        self.output_tex = output_tex
        self.data = {}
        self.load_data()

    def load_data(self):
        wb = openpyxl.load_workbook(self.excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            section_data = [cell.value for row in ws.iter_rows(values_only=True) for cell in row]
            self.data[sheet.lower()] = section_data

    def validate_and_format_dob(self, dob):
        # Try to parse the date in various common formats
        try:
            # Check if it's already in DD-MM-YYYY format
            if re.match(r"\d{2}-\d{2}-\d{4}", dob):
                return dob
            # If it's in some other common format, attempt to reformat it
            possible_formats = ["%d/%m/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d %B %Y"]
            for fmt in possible_formats:
                try:
                    parsed_date = datetime.strptime(dob, fmt)
                    return parsed_date.strftime("%d-%m-%Y")
                except ValueError:
                    continue
            # If all else fails, return the DOB unchanged
            return "Invalid Date Format"
        except Exception as e:
            return "Invalid Date Format"

    def generate_profile(self):
        dob = self.data['profile'][5]
        formatted_dob = self.validate_and_format_dob(dob)

        profile = f"""
        %-------------------------------------------------------------------------------
        % PROFILE
        %-------------------------------------------------------------------------------
        \\photo[rectangle,edge,right]{{./resume/lttsLogo.png}}\\First Name{{{self.data['profile'][0]}}}
        \\Last Name{{{self.data['profile'][1]}}}
        \\Designation (Job Title){{{self.data['profile'][2]}}}
        \\Mobile No{{{self.data['profile'][3]}}}
        \\Official E-Mail Id{{{self.data['profile'][4]}}}
        \\DOB:{{{formatted_dob}}}
        \\Github Profile Link (Optional){{{self.data['profile'][6]}}}
        """
        return profile

    def generate_tex(self):
        with open(self.output_tex, 'w') as f:
            f.write(r"""
            %-------------------------------------------------------------------------------
            % CONFIGURATIONS
            %-------------------------------------------------------------------------------
            \documentclass[11pt, a4paper]{awesome-cv} 
            \geometry{left=1.4cm, top=.8cm, right=1.4cm, bottom=1.8cm, footskip=.5cm}
            \colorlet{awesome}{awesome-skyblue}
            \setbool{acvSectionColorHighlight}{true}
            \renewcommand{\acvHeaderSocialSep}{\quad\textbar\quad}
            """)

            f.write(self.generate_profile())
            f.write(r"""
            \input{resume/summary.tex}
            \input{resume/skills.tex}
            \input{resume/competencies.tex}
            \input{resume/experience.tex}
            \input{resume/education.tex}

            \end{document}
            """)

if __name__ == "__main__":
    excel_file = "data.xlsx"  # Replace with your Excel file path
    output_tex = "main_resume.tex"
    
    generator = ResumeGenerator(excel_file, output_tex)
    generator.generate_tex()
